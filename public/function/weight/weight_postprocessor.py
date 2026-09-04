from __future__ import annotations

import configparser
import math
import os
import tempfile
from collections import deque
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from statistics import mean, median
from typing import Iterable, Mapping, Optional, Sequence

try:
    from loguru import logger
except ImportError:  # Allows the postprocessor to run in lightweight export tools.
    import logging

    logger = logging.getLogger(__name__)


CAGE_HEADER = "鼠笼号"
WEIGHT_HEADER = "称重重量测量值(g)"
TIME_HEADER = "获取时间"
EPOCH_WEIGHT_COLUMN_KEY = "WM_weight_num"
DEFAULT_CONFIG_PATH = (
    Path(__file__).resolve().parents[3]
    / "config"
    / "monitor_datas_config.ini"
)


@dataclass(frozen=True)
class WeightPostprocessConfig:
    # 监控数据界面是否用拟合值替换称重列；原始数据库数据不会被修改。
    display_fitted_weight: bool = True
    event_window_points: int = 15
    minimum_group_points: int = 5
    minimum_body_weight_g: float = 5.0
    maximum_body_weight_g: float = 80.0
    initial_weight_match_ratio: float = 0.20
    initial_weight_match_min_g: float = 2.0
    event_outlier_ratio: float = 0.05
    event_outlier_min_g: float = 1.0
    smoothing_points: int = 5
    ema_alpha: float = 0.35
    large_change_ratio: float = 0.20
    large_change_min_g: float = 5.0
    large_change_confirm_points: int = 3
    large_change_window_minutes: float = 10.0
    decimal_places: int = 3
    output_suffix: str = "_称重拟合"


@dataclass(frozen=True)
class WeightPostprocessResult:
    success: bool
    output_path: Optional[str] = None
    processed_sheets: int = 0
    processed_cages: int = 0
    skipped_cages: int = 0
    warnings: tuple[str, ...] = ()
    error: str = ""

    @property
    def summary(self) -> str:
        if not self.success:
            return self.error or "称重后处理失败"
        if self.output_path is None:
            return self.warnings[0] if self.warnings else "称重后处理未启用"
        return (
            f"称重拟合完成：{self.processed_sheets} 个工作表，"
            f"{self.processed_cages} 个笼子"
            + (
                f"，{self.skipped_cages} 个笼子尚未建立空秤基线"
                if self.skipped_cages
                else ""
            )
        )


@dataclass(frozen=True)
class _BaselineMatch:
    baseline: float
    baseline_ready_index: int
    initial_weight: float


def load_weight_postprocess_config(
        config_path: Optional[os.PathLike | str] = None,
) -> WeightPostprocessConfig:
    parser = configparser.ConfigParser(interpolation=None)
    path = Path(config_path) if config_path is not None else DEFAULT_CONFIG_PATH
    if path.is_file():
        parser.read(path, encoding="utf-8-sig")

    section = "WEIGHT_POSTPROCESS"
    get = lambda key, fallback: parser.get(section, key, fallback=fallback)
    get_int = lambda key, fallback: parser.getint(section, key, fallback=fallback)
    get_float = lambda key, fallback: parser.getfloat(section, key, fallback=fallback)
    get_bool = lambda key, fallback: parser.getboolean(section, key, fallback=fallback)

    config = WeightPostprocessConfig(
        display_fitted_weight=get_bool("display_fitted_weight", True),
        event_window_points=max(2, get_int("event_window_points", 15)),
        minimum_group_points=max(1, get_int("minimum_group_points", 5)),
        minimum_body_weight_g=max(0.0, get_float("minimum_body_weight_g", 5.0)),
        maximum_body_weight_g=max(0.0, get_float("maximum_body_weight_g", 80.0)),
        initial_weight_match_ratio=max(
            0.0, get_float("initial_weight_match_ratio", 0.20)
        ),
        initial_weight_match_min_g=max(
            0.0, get_float("initial_weight_match_min_g", 2.0)
        ),
        event_outlier_ratio=max(0.0, get_float("event_outlier_ratio", 0.05)),
        event_outlier_min_g=max(0.0, get_float("event_outlier_min_g", 1.0)),
        smoothing_points=max(1, get_int("smoothing_points", 5)),
        ema_alpha=min(1.0, max(0.0, get_float("ema_alpha", 0.35))),
        large_change_ratio=max(0.0, get_float("large_change_ratio", 0.20)),
        large_change_min_g=max(0.0, get_float("large_change_min_g", 5.0)),
        large_change_confirm_points=max(
            2, get_int("large_change_confirm_points", 3)
        ),
        large_change_window_minutes=max(
            0.1, get_float("large_change_window_minutes", 10.0)
        ),
        decimal_places=max(0, get_int("decimal_places", 3)),
        output_suffix=get("output_suffix", "_称重拟合").strip() or "_称重拟合",
    )
    if config.maximum_body_weight_g <= config.minimum_body_weight_g:
        raise ValueError("maximum_body_weight_g must be greater than minimum_body_weight_g")
    if config.minimum_group_points * 2 > config.event_window_points:
        raise ValueError("minimum_group_points cannot exceed half of event_window_points")
    return config


def _to_finite_float(value) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped.lower() in {"none", "null", "nan"}:
            return None
        value = stripped
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def _to_timestamp(value) -> Optional[float]:
    """Convert Excel/SQLite time values to seconds for time-window fitting."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, datetime):
        return value.timestamp()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).timestamp()
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            parsed = datetime.fromisoformat(stripped.replace("Z", "+00:00"))
        except ValueError:
            return None
        return parsed.timestamp()
    return _to_finite_float(value)


def _tolerance(reference: float, ratio: float, minimum_g: float) -> float:
    return max(minimum_g, abs(reference) * ratio)


def _robust_group_level(
        group: Sequence[tuple[int, float]],
        config: WeightPostprocessConfig,
) -> Optional[float]:
    values = [value for _, value in group]
    center = float(median(values))
    tolerance = _tolerance(
        center,
        config.event_outlier_ratio,
        config.event_outlier_min_g,
    )
    inliers = [value for value in values if abs(value - center) <= tolerance]
    if len(inliers) < config.minimum_group_points:
        return None
    return float(mean(inliers))


def _find_initial_baseline(
        values: Sequence[Optional[float]],
        initial_weight: Optional[float],
        config: WeightPostprocessConfig,
) -> Optional[_BaselineMatch]:
    """Find a fixed baseline from a rolling valid-point window.

    With a manual weight, the candidate must agree with that reference. Without
    one, the most stable chronological low/high split is selected automatically.
    """
    valid_buffer: list[tuple[int, float]] = []
    initial_tolerance = (
        _tolerance(
            initial_weight,
            config.initial_weight_match_ratio,
            config.initial_weight_match_min_g,
        )
        if initial_weight is not None
        else None
    )
    for index, value in enumerate(values):
        if value is None:
            continue
        valid_buffer.append((index, value))
        if len(valid_buffer) > config.event_window_points:
            valid_buffer.pop(0)
        if len(valid_buffer) < config.event_window_points:
            continue

        ranked = sorted(valid_buffer, key=lambda item: item[1])
        # score, baseline, candidate body weight
        candidates: list[tuple[float, float, float]] = []
        minimum_split = config.minimum_group_points
        maximum_split = config.event_window_points - config.minimum_group_points
        for split in range(minimum_split, maximum_split + 1):
            low_group = ranked[:split]
            high_group = ranked[split:]
            low_level = _robust_group_level(low_group, config)
            high_level = _robust_group_level(high_group, config)
            if low_level is None or high_level is None:
                continue
            candidate_weight = high_level - low_level
            if not (
                    config.minimum_body_weight_g
                    <= candidate_weight
                    <= config.maximum_body_weight_g
            ):
                continue
            low_spread = median(abs(value - low_level) for _, value in low_group)
            high_spread = median(abs(value - high_level) for _, value in high_group)
            stability_score = float(low_spread + high_spread)
            selection_score = (
                abs(candidate_weight - initial_weight)
                if initial_weight is not None
                else stability_score
            )
            candidates.append(
                (
                    selection_score,
                    low_level,
                    candidate_weight,
                )
            )

        if candidates:
            difference, baseline, candidate_weight = min(
                candidates,
                key=lambda item: item[0],
            )
            if initial_tolerance is None or difference <= initial_tolerance:
                return _BaselineMatch(
                    baseline=baseline,
                    baseline_ready_index=index,
                    initial_weight=candidate_weight,
                )
    return None


def fit_weight_series(
        values: Iterable,
        config: Optional[WeightPostprocessConfig] = None,
        timestamps: Optional[Iterable] = None,
        initial_weight: Optional[float] = None,
) -> tuple[list[Optional[float]], int]:
    """Fit one cage with a fixed baseline and continuous robust smoothing.

    The returned count is the number of confirmed large changes. It is not a
    requirement for using the continuously smoothed values.
    """
    config = config or WeightPostprocessConfig()
    numeric_values = [_to_finite_float(value) for value in values]
    numeric_timestamps: Optional[list[float]] = None
    if timestamps is not None:
        timestamp_values = [_to_timestamp(value) for value in timestamps]
        if len(timestamp_values) == len(numeric_values) and all(
                value is not None for value in timestamp_values
        ):
            numeric_timestamps = [float(value) for value in timestamp_values]
    initial_weight_value = _to_finite_float(initial_weight)
    if initial_weight_value is not None and not (
            config.minimum_body_weight_g
            <= initial_weight_value
            <= config.maximum_body_weight_g
    ):
        initial_weight_value = None
    baseline_match = _find_initial_baseline(numeric_values, initial_weight_value, config)
    if baseline_match is None:
        # 基线未建立前不输出人工体重或原始值，避免把尚未校正的数据伪装成体重。
        return [None] * len(numeric_values), 0

    fitted: list[Optional[float]] = [None for _ in numeric_values]
    smoothed_weight = baseline_match.initial_weight
    reference_weight = baseline_match.initial_weight
    recent_candidates = deque(maxlen=max(1, config.smoothing_points))
    pending: list[tuple[int, float]] = []
    confirmed_changes = 0
    confirm_window_seconds = config.large_change_window_minutes * 60.0

    for index, raw_value in enumerate(numeric_values):
        if index < baseline_match.baseline_ready_index:
            fitted[index] = None
            continue

        # A missing sample is an interruption, not a zero-weight measurement.
        # Do not allow samples on the two sides of a gap to confirm a change.
        if raw_value is None:
            pending.clear()
            fitted[index] = round(smoothed_weight, config.decimal_places)
            continue

        candidate_weight = raw_value - baseline_match.baseline
        if not (
                config.minimum_body_weight_g
                <= candidate_weight
                <= config.maximum_body_weight_g
        ):
            pending.clear()
            fitted[index] = round(smoothed_weight, config.decimal_places)
            continue

        large_change_limit = _tolerance(
            reference_weight,
            config.large_change_ratio,
            config.large_change_min_g,
        )
        if abs(candidate_weight - reference_weight) <= large_change_limit:
            pending.clear()
            recent_candidates.append(candidate_weight)
            smooth_target = float(median(recent_candidates))
            smoothed_weight = (
                config.ema_alpha * smooth_target
                + (1.0 - config.ema_alpha) * smoothed_weight
            )
            reference_weight = smooth_target
            fitted[index] = round(smoothed_weight, config.decimal_places)
            continue

        current_timestamp = (
            numeric_timestamps[index]
            if numeric_timestamps is not None
            else None
        )
        if pending and current_timestamp is not None:
            first_timestamp = numeric_timestamps[pending[0][0]]
            if current_timestamp - first_timestamp > confirm_window_seconds:
                pending.clear()

        if pending:
            pending_center = float(median(value for _, value in pending))
            pending_limit = _tolerance(
                pending_center,
                config.event_outlier_ratio,
                config.event_outlier_min_g,
            )
            if abs(candidate_weight - pending_center) <= pending_limit:
                pending.append((index, candidate_weight))
            else:
                pending[:] = [(index, candidate_weight)]
        else:
            pending.append((index, candidate_weight))

        if len(pending) >= config.large_change_confirm_points:
            confirmed_weight = float(
                median(
                    value
                    for _, value in pending[-config.large_change_confirm_points:]
                )
            )
            reference_weight = confirmed_weight
            recent_candidates.clear()
            recent_candidates.append(confirmed_weight)
            smoothed_weight = (
                config.ema_alpha * confirmed_weight
                + (1.0 - config.ema_alpha) * smoothed_weight
            )
            pending.clear()
            confirmed_changes += 1

        fitted[index] = round(smoothed_weight, config.decimal_places)

    return fitted, confirmed_changes


def _default_output_path(raw_excel_path: Path, suffix: str) -> Path:
    return raw_excel_path.with_name(f"{raw_excel_path.stem}{suffix}{raw_excel_path.suffix}")


def _initial_weight_for_cage(
        initial_weights: Optional[Mapping],
        cage_name: str,
        config: WeightPostprocessConfig,
) -> Optional[float]:
    if not initial_weights:
        return None
    normalized_cage_name = str(cage_name).strip()
    for key, value in initial_weights.items():
        if str(key).strip() != normalized_cage_name:
            continue
        numeric_value = _to_finite_float(value)
        if numeric_value is not None and (
                config.minimum_body_weight_g
                <= numeric_value
                <= config.maximum_body_weight_g
        ):
            return numeric_value
        return None
    return None


def weight_row_key(row: Mapping, fallback_cage=None) -> tuple[str, str | None]:
    """Return a stable key for matching a fitted history row to a page row."""
    cage = row.get("mouse_cage_number", fallback_cage)
    row_id = row.get("id")
    if row_id is None:
        row_id = row.get("time")
    return (
        "" if cage is None else str(cage).strip(),
        None if row_id is None else str(row_id).strip(),
    )


def fit_epoch_weight_rows(
        rows: Sequence[Mapping],
        initial_weights: Optional[Mapping] = None,
        config: Optional[WeightPostprocessConfig] = None,
        fallback_cage=None,
) -> dict[tuple[str, str | None], Optional[float]]:
    """Fit all cage groups in Epoch rows and return values keyed by row identity."""
    config = config or WeightPostprocessConfig()
    grouped: dict[str, list[tuple[int, Mapping]]] = {}
    for index, row in enumerate(rows or []):
        if EPOCH_WEIGHT_COLUMN_KEY not in row:
            continue
        cage = row.get("mouse_cage_number", fallback_cage)
        cage_key = "" if cage is None else str(cage).strip()
        grouped.setdefault(cage_key, []).append((index, row))

    fitted_by_key: dict[tuple[str, str | None], Optional[float]] = {}
    for cage_key, cage_rows in grouped.items():
        timestamped_rows = [
            (index, row, _to_timestamp(row.get("time")))
            for index, row in cage_rows
        ]
        has_complete_timestamps = bool(timestamped_rows) and all(
            timestamp is not None for _, _, timestamp in timestamped_rows
        )
        ordered_rows = (
            sorted(timestamped_rows, key=lambda item: float(item[2]))
            if has_complete_timestamps
            else timestamped_rows
        )
        values = [row.get(EPOCH_WEIGHT_COLUMN_KEY) for _, row, _ in ordered_rows]
        timestamps = (
            [timestamp for _, _, timestamp in ordered_rows]
            if has_complete_timestamps
            else None
        )
        fitted_values, _ = fit_weight_series(
            values,
            config=config,
            timestamps=timestamps,
            initial_weight=_initial_weight_for_cage(
                initial_weights,
                cage_key,
                config,
            ),
        )
        for (_, row, _), fitted_value in zip(ordered_rows, fitted_values):
            fitted_by_key[weight_row_key(row, fallback_cage=fallback_cage)] = fitted_value
    return fitted_by_key


def create_fitted_workbook(
        raw_excel_path: os.PathLike | str,
        output_path: Optional[os.PathLike | str] = None,
        config_path: Optional[os.PathLike | str] = None,
        initial_weights: Optional[Mapping] = None,
) -> WeightPostprocessResult:
    raw_path = Path(raw_excel_path).resolve()
    temp_path: Optional[Path] = None
    try:
        config = load_weight_postprocess_config(config_path)
        if not raw_path.is_file():
            raise FileNotFoundError(f"原始 Excel 不存在: {raw_path}")

        fitted_path = (
            Path(output_path).resolve()
            if output_path is not None
            else _default_output_path(raw_path, config.output_suffix)
        )
        if fitted_path == raw_path:
            raise ValueError("拟合结果不能覆盖原始 Excel")

        from openpyxl import load_workbook

        workbook = load_workbook(raw_path)
        matched_sheets = 0
        processed_sheets = 0
        processed_cage_names: set[str] = set()
        skipped_cage_names: set[str] = set()
        warnings: list[str] = []

        for worksheet in workbook.worksheets:
            headers = {
                str(cell.value).strip(): cell.column
                for cell in worksheet[1]
                if cell.value is not None
            }
            cage_column = headers.get(CAGE_HEADER)
            weight_column = headers.get(WEIGHT_HEADER)
            if cage_column is None or weight_column is None:
                continue
            matched_sheets += 1
            time_column = headers.get(TIME_HEADER)

            grouped_rows: dict[str, list[int]] = {}
            for row_number in range(2, worksheet.max_row + 1):
                cage_value = worksheet.cell(row_number, cage_column).value
                if cage_value is None or str(cage_value).strip() == "":
                    continue
                grouped_rows.setdefault(str(cage_value).strip(), []).append(row_number)

            sheet_processed = False
            for cage_name, row_numbers in grouped_rows.items():
                records = [
                    (
                        row_number,
                        worksheet.cell(row_number, weight_column).value,
                        _to_timestamp(
                            worksheet.cell(row_number, time_column).value
                        ) if time_column is not None else None,
                    )
                    for row_number in row_numbers
                ]
                has_complete_timestamps = (
                    time_column is not None
                    and bool(records)
                    and all(record[2] is not None for record in records)
                )
                if has_complete_timestamps:
                    records.sort(key=lambda record: float(record[2]))
                ordered_rows = [record[0] for record in records]
                raw_values = [record[1] for record in records]
                timestamps = (
                    [float(record[2]) for record in records]
                    if has_complete_timestamps
                    else None
                )
                numeric_count = sum(_to_finite_float(value) is not None for value in raw_values)
                initial_weight = _initial_weight_for_cage(
                    initial_weights,
                    cage_name,
                    config,
                )
                if numeric_count == 0:
                    for row_number in ordered_rows:
                        worksheet.cell(row_number, weight_column).value = None
                    continue

                # A cage can have a valid continuously smoothed series even
                # when no large-change event was confirmed. Without a manual
                # weight, the fitter searches for the baseline automatically.
                fitted_values, _ = fit_weight_series(
                    raw_values,
                    config,
                    timestamps=timestamps,
                    initial_weight=initial_weight,
                )
                if any(value is not None for value in fitted_values):
                    processed_cage_names.add(cage_name)
                else:
                    skipped_cage_names.add(cage_name)
                    warnings.append(
                        f"{worksheet.title}/{cage_name}: 尚未建立空秤基线，拟合列保留为 None"
                    )

                for row_number, fitted_value in zip(ordered_rows, fitted_values):
                    worksheet.cell(row_number, weight_column).value = (
                        None
                        if fitted_value is None
                        else round(float(fitted_value), config.decimal_places)
                    )
                sheet_processed = True

            if sheet_processed:
                processed_sheets += 1

        if matched_sheets == 0:
            raise ValueError(f"Excel 中未找到 {CAGE_HEADER} 和 {WEIGHT_HEADER} 数据列")
        if processed_sheets == 0:
            workbook.close()
            return WeightPostprocessResult(
                success=True,
                warnings=("未检测到有效称重数据，未生成称重拟合 Excel",),
            )

        fitted_path.parent.mkdir(parents=True, exist_ok=True)
        handle, temp_name = tempfile.mkstemp(
            prefix=f".{fitted_path.stem}.",
            suffix=fitted_path.suffix or ".xlsx",
            dir=fitted_path.parent,
        )
        os.close(handle)
        temp_path = Path(temp_name)
        workbook.save(temp_path)
        workbook.close()
        os.replace(temp_path, fitted_path)
        temp_path = None

        for warning in warnings:
            logger.warning(f"weight postprocess: {warning}")
        result = WeightPostprocessResult(
            success=True,
            output_path=str(fitted_path),
            processed_sheets=processed_sheets,
            processed_cages=len(processed_cage_names),
            skipped_cages=len(skipped_cage_names - processed_cage_names),
            warnings=tuple(warnings),
        )
        logger.info(f"{result.summary}: {fitted_path}")
        return result
    except Exception as error:
        if temp_path is not None and temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass
        logger.exception(f"称重后处理失败，原始 Excel 保持不变: {error}")
        return WeightPostprocessResult(success=False, error=str(error))
