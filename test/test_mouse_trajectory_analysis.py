from datetime import datetime
from pathlib import Path
import tempfile
import unittest

import numpy as np
from openpyxl import Workbook

from Module.mouse_trajectory_analysis.analysis_core import (
    ChannelTrajectory,
    ExperimentAnalysis,
    aggregate_distance,
    load_experiment_data,
    load_experiment_workbook,
    load_monitor_comparison_workbook,
    load_trajectory_experiment,
    find_matching_monitor_workbook,
    parse_experiment_time,
    scan_experiment_files,
    scan_trajectory_experiments,
    sleep_state_matrix,
)


class MouseTrajectoryAnalysisTests(unittest.TestCase):
    def test_original_monitor_workbook_is_preferred_over_fitted_workbook(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            original = root / "1_2026_08_01_12_00_02_000.xlsx"
            fitted = root / "1_2026_08_01_12_00_00_000_称重拟合.xlsx"
            original.touch()
            fitted.touch()

            selected = find_matching_monitor_workbook(
                root,
                datetime(2026, 8, 1, 12, 0, 0),
            )

            self.assertEqual(selected, original.resolve())

    def test_monitor_workbook_loads_the_fixed_comparison_fields(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "1_2026_08_01_12_00_00_000.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "每轮数据监控数据_通道1"
            sheet.append(
                [
                    "获取时间",
                    "称重重量测量值(g)",
                    "食物重量测量值(g)",
                    "温度测量值(°C)",
                    "氧浓度(%)",
                    "对齐后CO2",
                ]
            )
            sheet.append(
                [
                    "2026-08-01 12:00:01.000",
                    22.1,
                    3.2,
                    36.8,
                    20.9,
                    450.0,
                ]
            )
            sheet.append(
                [
                    "2026-08-01 12:00:02.000",
                    22.2,
                    3.4,
                    36.9,
                    21.0,
                    451.0,
                ]
            )
            workbook.save(path)

            start_timestamp = datetime(2026, 8, 1, 12, 0, 0).timestamp()
            series = load_monitor_comparison_workbook(path, start_timestamp)

            self.assertEqual(set(series[1]), {"weight", "food", "temperature", "oxygen", "co2"})
            np.testing.assert_allclose(
                series[1]["temperature"].elapsed_seconds,
                np.asarray([1.0, 2.0]),
            )
            np.testing.assert_allclose(
                series[1]["co2"].values,
                np.asarray([450.0, 451.0]),
            )

    def test_scan_uses_experiment_time_and_sorts_latest_first(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            older = root / "1_2026_07_27_10_00_00_001.xlsx"
            newer = root / "2_2026_07_28_09_30_00_002.xlsx"
            ignored = root / "~$2_2026_07_29_09_30_00_002.xlsx"
            for path in (older, newer, ignored):
                path.touch()

            records = scan_experiment_files(root)

            self.assertEqual([item.path.name for item in records], [newer.name, older.name])
            self.assertEqual(records[0].experiment_name, "2")
            self.assertEqual(
                parse_experiment_time(newer),
                datetime(2026, 7, 28, 9, 30, 0, 2000),
            )

    def test_workbook_distance_does_not_bridge_missing_frames(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "1_2026_08_01_12_00_00_000.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "目标检测结果监控数据_通道1"
            sheet.append(
                [
                    "序号",
                    "图像名称",
                    "深度图像是否存在",
                    "中心X",
                    "中心Y",
                    "获取时间",
                ]
            )
            rows = (
                (1, "f1", 1, 0.000, 0.000, "2026-08-01 12:00:00.000"),
                (2, "f2", 1, 0.001, 0.000, "2026-08-01 12:00:01.000"),
                (3, "f3", 0, None, None, "2026-08-01 12:00:02.000"),
                (4, "f4", 1, 0.003, 0.000, "2026-08-01 12:00:03.000"),
                (5, "f5", 1, 0.004, 0.000, "2026-08-01 12:00:04.000"),
            )
            for row in rows:
                sheet.append(row)
            workbook.save(path)

            analysis = load_experiment_workbook(path)
            channel = analysis.channels[1]

            np.testing.assert_allclose(
                channel.distance_steps_mm,
                np.asarray([0.0, 1.0, 0.0, 0.0, 1.0]),
            )
            self.assertAlmostEqual(channel.total_distance_mm, 2.0)
            self.assertEqual(channel.total_rows, 5)
            self.assertEqual(channel.valid_rows, 4)
            _, per_second = aggregate_distance(channel, 1.0)
            self.assertAlmostEqual(float(per_second.sum()), channel.total_distance_mm)

    def test_sleep_starts_from_the_fourth_consecutive_blank_segment(self):
        times = np.arange(40, dtype=float)
        invalid = np.zeros(40, dtype=bool)
        zeros = np.zeros(40, dtype=float)
        nan_values = np.full(40, np.nan, dtype=float)
        channel = ChannelTrajectory(
            channel=1,
            times=times,
            elapsed_seconds=times,
            x_mm=nan_values.copy(),
            y_mm=nan_values.copy(),
            valid=invalid,
            distance_steps_mm=zeros.copy(),
            cumulative_distance_mm=zeros.copy(),
            gap_limit_seconds=5.0,
            source_sheets=("目标检测结果监控数据_通道1",),
        )
        empty = ChannelTrajectory(
            channel=2,
            times=np.asarray([], dtype=float),
            elapsed_seconds=np.asarray([], dtype=float),
            x_mm=np.asarray([], dtype=float),
            y_mm=np.asarray([], dtype=float),
            valid=np.asarray([], dtype=bool),
            distance_steps_mm=np.asarray([], dtype=float),
            cumulative_distance_mm=np.asarray([], dtype=float),
            gap_limit_seconds=5.0,
            source_sheets=(),
        )
        channels = {1: channel}
        for channel_number in range(2, 9):
            channels[channel_number] = ChannelTrajectory(
                channel=channel_number,
                times=empty.times.copy(),
                elapsed_seconds=empty.elapsed_seconds.copy(),
                x_mm=empty.x_mm.copy(),
                y_mm=empty.y_mm.copy(),
                valid=empty.valid.copy(),
                distance_steps_mm=empty.distance_steps_mm.copy(),
                cumulative_distance_mm=empty.cumulative_distance_mm.copy(),
                gap_limit_seconds=5.0,
                source_sheets=(),
            )
        analysis = ExperimentAnalysis(
            source_path=Path("test.xlsx"),
            started_at=datetime.fromtimestamp(0),
            start_timestamp=0.0,
            end_timestamp=40.0,
            channels=channels,
        )

        states, _ = sleep_state_matrix(analysis, movement_threshold_mm=5.0)

        np.testing.assert_allclose(states[0], np.asarray([0.0, 0.0, 0.0, 1.0]))
        self.assertTrue(np.all(np.isnan(states[1:])))

    def test_empty_excel_uses_time_matched_trajectory_csv(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            excel_path = root / "monitor" / "1_2026_08_07_11_03_15_626.xlsx"
            excel_path.parent.mkdir()
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "目标检测结果监控数据_通道1"
            sheet.append(["序号", "中心X", "中心Y", "获取时间"])
            workbook.save(excel_path)

            export_path = root / "trajectory" / "experiment_20260807_110313_461_pid8272"
            csv_dir = export_path / "cage_1" / "data"
            csv_dir.mkdir(parents=True)
            (csv_dir / "trajectory.csv").write_text(
                "timestamp,datetime,status,X,Y\n"
                "1786071798.0,2026-08-07 11:03:18.000,ok,0,0\n"
                "1786071799.0,2026-08-07 11:03:19.000,no_mouse,,\n"
                "1786071800.0,2026-08-07 11:03:20.000,ok,3,4\n"
                "1786071801.0,2026-08-07 11:03:21.000,ok,6,8\n",
                encoding="utf-8",
            )

            analysis = load_experiment_data(
                excel_path,
                trajectory_root=export_path.parent,
            )

            self.assertEqual(analysis.coordinate_source_path, export_path)
            self.assertEqual(analysis.channels[1].valid_rows, 3)
            self.assertAlmostEqual(analysis.channels[1].total_distance_mm, 5.0)

    def test_scan_and_load_trajectory_experiment_directly(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            older = root / "experiment_20260806_164545_650_pid23860"
            newer = root / "experiment_20260807_110313_461_pid8272"
            ignored = root / "experiment_20260808_120000_000_pid9999"
            for export_path in (older, newer):
                csv_dir = export_path / "cage_1" / "data"
                csv_dir.mkdir(parents=True)
                (csv_dir / "trajectory.csv").write_text(
                    "timestamp,datetime,status,X,Y\n"
                    "1786071798.0,2026-08-07 11:03:18.000,ok,0,0\n"
                    "1786071799.0,2026-08-07 11:03:19.000,no_mouse,,\n"
                    "1786071800.0,2026-08-07 11:03:20.000,ok,3,4\n"
                    "1786071801.0,2026-08-07 11:03:21.000,ok,6,8\n",
                    encoding="utf-8",
                )
            ignored.mkdir()

            records = scan_trajectory_experiments(root)

            self.assertEqual([record.path for record in records], [newer, older])
            analysis = load_trajectory_experiment(newer)
            self.assertEqual(analysis.source_path, newer)
            self.assertEqual(analysis.coordinate_source_path, newer)
            self.assertEqual(analysis.channels[1].total_rows, 4)
            self.assertEqual(analysis.channels[1].valid_rows, 3)
            self.assertAlmostEqual(analysis.channels[1].total_distance_mm, 5.0)

    def test_direct_export_with_only_blank_frames_is_still_analyzable(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            export_path = Path(temp_dir) / "experiment_20260807_120000_000_pid1"
            csv_dir = export_path / "cage_1" / "data"
            csv_dir.mkdir(parents=True)
            (csv_dir / "trajectory.csv").write_text(
                "timestamp,datetime,status,X,Y\n"
                "1786075200.0,2026-08-07 12:00:00.000,no_mouse,,\n"
                "1786075201.0,2026-08-07 12:00:01.000,no_mouse,,\n",
                encoding="utf-8",
            )

            analysis = load_trajectory_experiment(export_path)

            self.assertEqual(analysis.channels[1].total_rows, 2)
            self.assertEqual(analysis.channels[1].valid_rows, 0)
            self.assertEqual(analysis.channels[1].total_distance_mm, 0.0)


if __name__ == "__main__":
    unittest.main()
