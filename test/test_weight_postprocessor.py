import tempfile
import unittest
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

from public.function.weight.weight_postprocessor import (
    CAGE_HEADER,
    TIME_HEADER,
    WEIGHT_HEADER,
    WeightPostprocessConfig,
    create_fitted_workbook,
    fit_weight_series,
)


def _weighing_cycle(baseline, weight, *, high_values=None):
    """Make 15-point baseline setup followed by one weighing event."""
    low_values = [baseline - 0.1, baseline, baseline + 0.1, baseline] * 2
    if high_values is None:
        high_values = [baseline + weight - 0.1, baseline + weight, baseline + weight + 0.1] * 4
    return low_values + high_values + [baseline, baseline + 0.1, baseline - 0.1] * 3


class WeightPostprocessorTest(unittest.TestCase):
    def test_manual_weight_establishes_baseline_and_event_updates_after_window(self):
        values = _weighing_cycle(-10.0, 46.2)

        fitted, event_count = fit_weight_series(values, initial_weight=47.0)

        self.assertEqual(1, event_count)
        self.assertEqual(47.0, fitted[0])
        self.assertEqual(47.0, fitted[21])
        self.assertAlmostEqual(46.2, fitted[22], delta=0.1)
        self.assertAlmostEqual(46.2, fitted[-1], delta=0.1)

    def test_negative_raw_readings_can_use_a_negative_empty_scale_baseline(self):
        values = _weighing_cycle(-20.0, 26.1)

        fitted, event_count = fit_weight_series(values, initial_weight=26.0)

        self.assertEqual(1, event_count)
        self.assertAlmostEqual(26.1, fitted[-1], places=2)

    def test_event_filter_discards_a_single_high_outlier(self):
        high_values = [47.0, 46.9, 47.1, 47.0, 75.0, 46.9, 47.1, 47.0, 47.0, 46.9]
        values = _weighing_cycle(0.0, 47.0, high_values=high_values)

        fitted, event_count = fit_weight_series(values, initial_weight=47.0)

        self.assertEqual(1, event_count)
        self.assertEqual(fitted[22], fitted[-1])
        self.assertAlmostEqual(47.0, fitted[-1], delta=0.1)
        self.assertNotIn(75.0, fitted)

    def test_rejected_event_does_not_replace_last_confirmed_weight(self):
        first_event = _weighing_cycle(0.0, 47.0)
        second_event = [30.0, 29.9, 30.1] * 4 + [0.0, 0.1, -0.1] * 3
        values = first_event + [0.0] * 6 + second_event

        fitted, event_count = fit_weight_series(values, initial_weight=47.0)

        self.assertEqual(1, event_count)
        self.assertAlmostEqual(47.0, fitted[-1], delta=0.1)

    def test_manual_weight_is_held_when_no_baseline_or_event_is_available(self):
        fitted, event_count = fit_weight_series(
            [None, 0.0, 0.2, None] * 5,
            initial_weight=23.5,
        )

        self.assertEqual(0, event_count)
        self.assertEqual([23.5] * 20, fitted)

    def test_missing_manual_weight_has_no_fitted_value(self):
        fitted, event_count = fit_weight_series(_weighing_cycle(0.0, 25.0))

        self.assertEqual(0, event_count)
        self.assertTrue(all(value is None for value in fitted))

    def test_workbook_keeps_raw_file_and_uses_initial_weight_for_every_cage(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            raw_path = Path(temp_dir) / "experiment.xlsx"
            fitted_path = Path(temp_dir) / "experiment_称重拟合.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "每轮数据监控数据"
            sheet.append(["序号", CAGE_HEADER, WEIGHT_HEADER, "备注"])

            cage_one_values = _weighing_cycle(0.0, 47.0)
            cage_one_values[12] = 75.0
            for index, value in enumerate(cage_one_values, start=1):
                sheet.append([index, 1, value, f"原备注{index}"])
            for index in range(1, 5):
                sheet.append([100 + index, 2, None, "笼2"])

            raw_weight_sheet = workbook.create_sheet("称重模块监控数据_通道1")
            raw_weight_sheet.append(["序号", "重量测量值(g)", "备注"])
            raw_weight_sheet.append([1, 75.0, "原始异常值"])
            workbook.save(raw_path)

            result = create_fitted_workbook(
                raw_path,
                output_path=fitted_path,
                initial_weights={"1": 47.0, "2": 33.0},
            )

            self.assertTrue(result.success, result.error)
            self.assertTrue(raw_path.exists())
            self.assertTrue(fitted_path.exists())

            raw_workbook = load_workbook(raw_path, data_only=True)
            fitted_workbook = load_workbook(fitted_path, data_only=True)
            raw_sheet = raw_workbook["每轮数据监控数据"]
            fitted_sheet = fitted_workbook["每轮数据监控数据"]
            cage_one_count = len(cage_one_values)
            raw_values = [
                raw_sheet.cell(row, 3).value
                for row in range(2, 2 + cage_one_count + 4)
            ]
            fitted_values = [
                fitted_sheet.cell(row, 3).value
                for row in range(2, 2 + cage_one_count + 4)
            ]

            self.assertIn(75.0, raw_values)
            self.assertNotIn(75.0, fitted_values)
            self.assertTrue(all(45.0 <= value <= 48.0 for value in fitted_values[:cage_one_count]))
            self.assertEqual([33.0] * 4, fitted_values[cage_one_count:])
            self.assertEqual("原备注1", fitted_sheet.cell(2, 4).value)
            self.assertEqual(75.0, fitted_workbook["称重模块监控数据_通道1"].cell(2, 2).value)

    def test_workbook_sorts_by_time_for_fit_then_writes_to_original_rows(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            raw_path = Path(temp_dir) / "reversed.xlsx"
            fitted_path = Path(temp_dir) / "reversed_称重拟合.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([CAGE_HEADER, WEIGHT_HEADER, TIME_HEADER])

            base_time = datetime(2026, 8, 24, 12, 0, 0)
            chronological = _weighing_cycle(100.0, 18.0)
            for offset, value in reversed(list(enumerate(chronological))):
                sheet.append([1, value, base_time + timedelta(seconds=offset)])
            workbook.save(raw_path)

            result = create_fitted_workbook(
                raw_path,
                output_path=fitted_path,
                initial_weights={"1": 18.0},
            )

            self.assertTrue(result.success, result.error)
            fitted_workbook = load_workbook(fitted_path, data_only=True)
            fitted_sheet = fitted_workbook.active
            fitted_by_time = {
                fitted_sheet.cell(row, 3).value: fitted_sheet.cell(row, 2).value
                for row in range(2, fitted_sheet.max_row + 1)
            }
            self.assertEqual(18.0, fitted_by_time[base_time])
            self.assertAlmostEqual(
                18.0,
                fitted_by_time[base_time + timedelta(seconds=len(chronological) - 1)],
                delta=0.1,
            )

    def test_no_weight_data_without_initial_weight_does_not_create_output(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            raw_path = Path(temp_dir) / "empty_weight.xlsx"
            fitted_path = Path(temp_dir) / "empty_weight_称重拟合.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([CAGE_HEADER, WEIGHT_HEADER])
            sheet.append([1, "None"])
            workbook.save(raw_path)

            result = create_fitted_workbook(raw_path, output_path=fitted_path)

            self.assertTrue(result.success)
            self.assertIsNone(result.output_path)
            self.assertFalse(fitted_path.exists())
            self.assertTrue(raw_path.exists())


if __name__ == "__main__":
    unittest.main()
