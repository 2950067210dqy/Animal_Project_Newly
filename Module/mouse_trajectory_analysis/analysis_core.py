from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
import csv
import math
from pathlib import Path
import re
from typing import Callable, Iterable

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


EXPERIMENT_TIME_PATTERN = re.compile(
    r"(?P<year>20\d{2})_(?P<month>\d{2})_(?P<day>\d{2})_"
    r"(?P<hour>\d{2})_(?P<minute>\d{2})_(?P<second>\d{2})_"
    r"(?P<millisecond>\d{3})(?!\d)"
)
CHANNEL_PATTERNS = (
    re.compile(r"(?:通道|笼子)\s*[_-]?\s*(\d+)", re.IGNORECASE),
    re.compile(r"(?:cage|channel)\s*[_-]?\s*(\d+)", re.IGNORECASE),
)
TRAJECTORY_SHEET_MARKERS = ("目标检测", "轨迹", "trajectory")
TRAJECTORY_EXPORT_TIME_PATTERN = re.compile(
    r"(?P<year>20\d{2})(?P<month>\d{2})(?P<day>\d{2})_"
    r"(?P<hour>\d{2})(?P<minute>\d{2})(?P<second>\d{2})_"
    r"(?P<millisecond>\d{3})(?!\d)"
)


@dataclass(frozen=True)
class ExperimentFile:
    path: Path
    started_at: datetime
    experiment_name: str

    @property
    def display_text(self) -> str:
        time_text = self.started_at.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        name = self.experiment_name.strip(" _-")
        return f"{time_text}  |  实验 {name}" if name else time_text


@dataclass
class ChannelTrajectory:
    channel: int
    times: np.ndarray
    elapsed_seconds: np.ndarray
    x_mm: np.ndarray
    y_mm: np.ndarray
    valid: np.ndarray
    distance_steps_mm: np.ndarray
    cumulative_distance_mm: np.ndarray
    gap_limit_seconds: float
    source_sheets: tuple[str, ...]

    @property
    def total_rows(self) -> int:
        return int(self.times.size)

    @property
    def valid_rows(self) -> int:
        return int(np.count_nonzero(self.valid))

    @property
    def detection_rate(self) -> float:
        if self.total_rows == 0:
            return 0.0
        return self.valid_rows / self.total_rows

    @property
    def total_distance_mm(self) -> float:
        if self.cumulative_distance_mm.size == 0:
            return 0.0
        return float(self.cumulative_distance_mm[-1])


@dataclass
class ExperimentAnalysis:
    source_path: Path
    started_at: datetime
    start_timestamp: float
    end_timestamp: float
    channels: dict[int, ChannelTrajectory]
    coordinate_source_path: Path | None = None

    @property
    def duration_seconds(self) -> float:
        return max(self.end_timestamp - self.start_timestamp, 0.0)


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[\s_()（）\[\]【】]+", "", str(value)).lower()


def parse_experiment_time(path: str | Path) -> datetime | None:
    match = EXPERIMENT_TIME_PATTERN.search(Path(path).stem)
    if match is None:
        return None
    values = {key: int(value) for key, value in match.groupdict().items()}
    return datetime(
        values["year"],
        values["month"],
        values["day"],
        values["hour"],
        values["minute"],
        values["second"],
        values["millisecond"] * 1000,
    )


def parse_trajectory_export_time(path: str | Path) -> datetime | None:
    match = TRAJECTORY_EXPORT_TIME_PATTERN.search(Path(path).name)
    if match is None:
        return None
    values = {key: int(value) for key, value in match.groupdict().items()}
    return datetime(
        values["year"],
        values["month"],
        values["day"],
        values["hour"],
        values["minute"],
        values["second"],
        values["millisecond"] * 1000,
    )


def _experiment_name(path: Path) -> str:
    match = EXPERIMENT_TIME_PATTERN.search(path.stem)
    if match is None:
        return path.stem
    return path.stem[: match.start()].rstrip(" _-")


def scan_experiment_files(root: str | Path) -> list[ExperimentFile]:
    root_path = Path(root).expanduser().resolve()
    if not root_path.exists():
        return []

    records: list[ExperimentFile] = []
    for path in root_path.rglob("*.xlsx"):
        if path.name.startswith(("~$", ".")) or not path.is_file():
            continue
        try:
            started_at = parse_experiment_time(path)
            if started_at is None:
                started_at = datetime.fromtimestamp(path.stat().st_mtime)
            records.append(
                ExperimentFile(
                    path=path.resolve(),
                    started_at=started_at,
                    experiment_name=_experiment_name(path),
                )
            )
        except OSError:
            continue
    records.sort(key=lambda item: (item.started_at, item.path.name), reverse=True)
    return records


def scan_trajectory_experiments(root: str | Path) -> list[ExperimentFile]:
    """Return trajectory export directories that contain at least one cage CSV."""
    root_path = Path(root).expanduser().resolve()
    if not root_path.exists():
        return []

    records: list[ExperimentFile] = []
    for path in root_path.iterdir():
        if not path.is_dir() or path.name.startswith("."):
            continue
        csv_paths = list(path.glob("cage_*/data/trajectory.csv"))
        if not csv_paths:
            continue
        try:
            started_at = parse_trajectory_export_time(path)
            if started_at is None:
                started_at = datetime.fromtimestamp(
                    min(csv_path.stat().st_mtime for csv_path in csv_paths)
                )
            records.append(
                ExperimentFile(
                    path=path.resolve(),
                    started_at=started_at,
                    experiment_name="",
                )
            )
        except OSError:
            continue
    records.sort(key=lambda item: (item.started_at, item.path.name), reverse=True)
    return records


def _extract_channel(sheet_name: str) -> int | None:
    for pattern in CHANNEL_PATTERNS:
        match = pattern.search(sheet_name)
        if match is not None:
            channel = int(match.group(1))
            return channel if 1 <= channel <= 8 else None
    return None


def _is_trajectory_sheet(sheet_name: str) -> bool:
    lower_name = sheet_name.lower()
    return any(marker.lower() in lower_name for marker in TRAJECTORY_SHEET_MARKERS)


def _find_column(header_map: dict[str, int], aliases: Iterable[str]) -> int | None:
    for alias in aliases:
        index = header_map.get(_normalize_header(alias))
        if index is not None:
            return index
    return None


def _coordinate_columns(header_map: dict[str, int]) -> tuple[int, int, str] | None:
    candidates = (
        (("topXmm", "xmm"), ("topYmm", "ymm"), "mm"),
        (("X",), ("Y",), "mm"),
        (("中心X", "center_x", "centerX"), ("中心Y", "center_y", "centerY"), "auto"),
        (("中位数X", "median_x", "medianX"), ("中位数Y", "median_y", "medianY"), "auto"),
    )
    for x_aliases, y_aliases, unit_hint in candidates:
        x_index = _find_column(header_map, x_aliases)
        y_index = _find_column(header_map, y_aliases)
        if x_index is not None and y_index is not None:
            return x_index, y_index, unit_hint
    return None


def _safe_float(value: object) -> float:
    if value is None or isinstance(value, bool):
        return math.nan
    try:
        result = float(value)
    except (TypeError, ValueError):
        return math.nan
    return result if math.isfinite(result) else math.nan


def _parse_datetime_text(value: str) -> float | None:
    text = value.strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("/", "-")).timestamp()
    except ValueError:
        pass

    match = EXPERIMENT_TIME_PATTERN.search(text)
    if match is not None:
        parsed = parse_experiment_time(text)
        return parsed.timestamp() if parsed is not None else None

    for date_format in (
        "%Y-%m-%d %H:%M:%S",
        "%Y_%m_%d_%H_%M_%S",
        "%Y%m%d_%H%M%S",
    ):
        try:
            return datetime.strptime(text, date_format).timestamp()
        except ValueError:
            continue
    return None


def _parse_timestamp(value: object) -> float | None:
    if isinstance(value, datetime):
        return value.timestamp()
    if isinstance(value, str):
        return _parse_datetime_text(value)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric = float(value)
        if not math.isfinite(numeric):
            return None
        if numeric >= 100_000_000:
            return numeric
        if 1 <= numeric <= 4_000_000:
            try:
                return from_excel(numeric).timestamp()
            except (OverflowError, ValueError, TypeError):
                return None
    return None


def _cell(row: tuple[object, ...], index: int | None) -> object:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def _status_allows_detection(value: object) -> bool:
    if value is None:
        return True
    normalized = str(value).strip().lower()
    if not normalized:
        return True
    return normalized in {"ok", "detected", "true", "1", "有效", "已检测", "检测到"}


def _depth_flag_allows_detection(value: object) -> bool:
    if value is None or str(value).strip() == "":
        return True
    if isinstance(value, str):
        return value.strip().lower() not in {"0", "false", "no", "否", "不存在"}
    return bool(value)


def _coordinate_scale(unit_hint: str, x_values: list[float], y_values: list[float]) -> float:
    if unit_hint == "mm":
        return 1.0
    finite_values = [abs(value) for value in (*x_values, *y_values) if math.isfinite(value)]
    if not finite_values:
        return 1.0
    percentile = float(np.percentile(np.asarray(finite_values, dtype=float), 95))
    return 1000.0 if percentile <= 5.0 else 1.0


def _estimate_gap_limit(times: np.ndarray) -> float:
    if times.size < 2:
        return 5.0
    intervals = np.diff(times)
    intervals = intervals[np.isfinite(intervals) & (intervals > 0)]
    if intervals.size == 0:
        return 5.0
    return max(5.0, float(np.median(intervals)) * 3.0)


def _build_channel(
    channel: int,
    rows: list[tuple[float, float, float]],
    start_timestamp: float,
    source_sheets: list[str],
) -> ChannelTrajectory:
    if not rows:
        empty = np.asarray([], dtype=float)
        return ChannelTrajectory(
            channel=channel,
            times=empty.copy(),
            elapsed_seconds=empty.copy(),
            x_mm=empty.copy(),
            y_mm=empty.copy(),
            valid=np.asarray([], dtype=bool),
            distance_steps_mm=empty.copy(),
            cumulative_distance_mm=empty.copy(),
            gap_limit_seconds=5.0,
            source_sheets=tuple(source_sheets),
        )

    rows.sort(key=lambda item: item[0])
    values = np.asarray(rows, dtype=float)
    times = values[:, 0]
    x_values = values[:, 1]
    y_values = values[:, 2]
    valid = np.isfinite(x_values) & np.isfinite(y_values)
    gap_limit = _estimate_gap_limit(times)
    distance_steps = np.zeros(times.size, dtype=float)

    if times.size > 1:
        time_steps = np.diff(times)
        consecutive = (
            valid[1:]
            & valid[:-1]
            & np.isfinite(time_steps)
            & (time_steps >= 0)
            & (time_steps <= gap_limit)
        )
        x_steps = np.diff(x_values)
        y_steps = np.diff(y_values)
        pair_distances = np.hypot(x_steps, y_steps)
        distance_steps[1:] = np.where(consecutive, pair_distances, 0.0)

    return ChannelTrajectory(
        channel=channel,
        times=times,
        elapsed_seconds=np.maximum(times - start_timestamp, 0.0),
        x_mm=x_values,
        y_mm=y_values,
        valid=valid,
        distance_steps_mm=distance_steps,
        cumulative_distance_mm=np.cumsum(distance_steps),
        gap_limit_seconds=gap_limit,
        source_sheets=tuple(source_sheets),
    )


def load_experiment_workbook(
    path: str | Path,
    interruption_requested: Callable[[], bool] | None = None,
) -> ExperimentAnalysis:
    source_path = Path(path).expanduser().resolve()
    parsed_start = parse_experiment_time(source_path)
    if parsed_start is None:
        parsed_start = datetime.fromtimestamp(source_path.stat().st_mtime)

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    raw_rows: dict[int, list[tuple[float, float, float]]] = {channel: [] for channel in range(1, 9)}
    source_sheets: dict[int, list[str]] = {channel: [] for channel in range(1, 9)}
    all_timestamps: list[float] = []

    try:
        for sheet_name in workbook.sheetnames:
            if interruption_requested is not None and interruption_requested():
                raise InterruptedError("Excel读取已取消")
            channel = _extract_channel(sheet_name)
            if channel is None or not _is_trajectory_sheet(sheet_name):
                continue

            worksheet = workbook[sheet_name]
            iterator = worksheet.iter_rows(values_only=True)
            header = next(iterator, None)
            if header is None:
                continue
            header_map = {
                _normalize_header(value): index
                for index, value in enumerate(header)
                if _normalize_header(value)
            }
            coordinate_columns = _coordinate_columns(header_map)
            if coordinate_columns is None:
                continue
            x_index, y_index, unit_hint = coordinate_columns
            time_index = _find_column(
                header_map,
                ("获取时间", "datetime", "timestamp", "captureTimestamp", "time"),
            )
            frame_name_index = _find_column(header_map, ("图像名称", "frameName", "imageName"))
            status_index = _find_column(header_map, ("status", "检测状态", "状态"))
            depth_flag_index = _find_column(
                header_map,
                ("深度图像是否存在", "isExistDepthImage", "is_exist_depth_image"),
            )

            pending_rows: list[tuple[float | None, float, float, bool]] = []
            x_samples: list[float] = []
            y_samples: list[float] = []
            for row_number, row in enumerate(iterator, start=2):
                if row_number % 500 == 0 and interruption_requested is not None and interruption_requested():
                    raise InterruptedError("Excel读取已取消")
                timestamp = _parse_timestamp(_cell(row, time_index))
                if timestamp is None:
                    timestamp = _parse_timestamp(_cell(row, frame_name_index))
                x_value = _safe_float(_cell(row, x_index))
                y_value = _safe_float(_cell(row, y_index))
                detection_allowed = _status_allows_detection(_cell(row, status_index))
                detection_allowed = detection_allowed and _depth_flag_allows_detection(
                    _cell(row, depth_flag_index)
                )
                if math.isfinite(x_value) and math.isfinite(y_value):
                    x_samples.append(x_value)
                    y_samples.append(y_value)
                pending_rows.append((timestamp, x_value, y_value, detection_allowed))

            if not pending_rows:
                source_sheets[channel].append(sheet_name)
                continue

            scale = _coordinate_scale(unit_hint, x_samples, y_samples)
            fallback_interval = 0.1
            last_timestamp: float | None = None
            for row_index, (timestamp, x_value, y_value, detection_allowed) in enumerate(pending_rows):
                if timestamp is None:
                    timestamp = (
                        last_timestamp + fallback_interval
                        if last_timestamp is not None
                        else parsed_start.timestamp() + row_index * fallback_interval
                    )
                last_timestamp = timestamp
                if not detection_allowed:
                    x_value = math.nan
                    y_value = math.nan
                elif math.isfinite(x_value) and math.isfinite(y_value):
                    x_value *= scale
                    y_value *= scale
                raw_rows[channel].append((timestamp, x_value, y_value))
                all_timestamps.append(timestamp)
            source_sheets[channel].append(sheet_name)
    finally:
        workbook.close()

    start_timestamp = parsed_start.timestamp()
    if all_timestamps:
        earliest = min(all_timestamps)
        if earliest < start_timestamp - 60.0:
            start_timestamp = earliest
        end_timestamp = max(max(all_timestamps), start_timestamp)
    else:
        end_timestamp = start_timestamp

    channels = {
        channel: _build_channel(
            channel,
            raw_rows[channel],
            start_timestamp,
            source_sheets[channel],
        )
        for channel in range(1, 9)
    }
    if not any(channel.total_rows for channel in channels.values()):
        raise ValueError("Excel中未找到可分析的通道轨迹数据")

    return ExperimentAnalysis(
        source_path=source_path,
        started_at=datetime.fromtimestamp(start_timestamp),
        start_timestamp=start_timestamp,
        end_timestamp=end_timestamp,
        channels=channels,
        coordinate_source_path=source_path,
    )


def find_matching_trajectory_export(
    trajectory_root: str | Path,
    experiment_started_at: datetime,
    maximum_difference_seconds: float = 300.0,
) -> Path | None:
    root = Path(trajectory_root).expanduser().resolve()
    if not root.exists():
        return None

    best_path: Path | None = None
    best_difference = float("inf")
    for candidate in root.iterdir():
        if not candidate.is_dir():
            continue
        candidate_time = parse_trajectory_export_time(candidate)
        if candidate_time is None:
            continue
        difference = abs((candidate_time - experiment_started_at).total_seconds())
        if difference > maximum_difference_seconds or difference >= best_difference:
            continue
        if not any(candidate.glob("cage_*/data/trajectory.csv")):
            continue
        best_path = candidate
        best_difference = difference
    return best_path


def _load_trajectory_csv_export(
    source_path: Path,
    export_path: Path,
    experiment_started_at: datetime,
    interruption_requested: Callable[[], bool] | None = None,
) -> ExperimentAnalysis:
    raw_rows: dict[int, list[tuple[float, float, float]]] = {
        channel: [] for channel in range(1, 9)
    }
    source_sheets: dict[int, list[str]] = {
        channel: [] for channel in range(1, 9)
    }
    all_timestamps: list[float] = []

    for channel in range(1, 9):
        csv_path = export_path / f"cage_{channel}" / "data" / "trajectory.csv"
        if not csv_path.exists():
            continue
        source_sheets[channel].append(str(csv_path))
        with csv_path.open("r", encoding="utf-8-sig", newline="") as csv_file:
            reader = csv.DictReader(csv_file)
            for row_number, row in enumerate(reader, start=1):
                if row_number % 500 == 0 and interruption_requested is not None and interruption_requested():
                    raise InterruptedError("轨迹数据读取已取消")
                timestamp = None
                for field_name in ("captureTimestamp", "timestamp", "datetime", "frameName"):
                    timestamp = _parse_timestamp(row.get(field_name))
                    if timestamp is not None:
                        break
                if timestamp is None:
                    timestamp = experiment_started_at.timestamp() + (row_number - 1) * 0.1

                status_allows = _status_allows_detection(row.get("status"))
                x_value = _safe_float(row.get("X"))
                y_value = _safe_float(row.get("Y"))
                if not status_allows:
                    x_value = math.nan
                    y_value = math.nan
                raw_rows[channel].append((timestamp, x_value, y_value))
                all_timestamps.append(timestamp)

    if not all_timestamps:
        raise ValueError("匹配到的轨迹目录中没有可读取的trajectory.csv数据")

    start_timestamp = min(experiment_started_at.timestamp(), min(all_timestamps))
    end_timestamp = max(start_timestamp, max(all_timestamps))
    channels = {
        channel: _build_channel(
            channel,
            raw_rows[channel],
            start_timestamp,
            source_sheets[channel],
        )
        for channel in range(1, 9)
    }
    if not any(channel.total_rows for channel in channels.values()):
        raise ValueError("匹配到的轨迹文件中没有有效坐标")

    return ExperimentAnalysis(
        source_path=source_path,
        started_at=datetime.fromtimestamp(start_timestamp),
        start_timestamp=start_timestamp,
        end_timestamp=end_timestamp,
        channels=channels,
        coordinate_source_path=export_path,
    )


def load_trajectory_experiment(
    export_path: str | Path,
    interruption_requested: Callable[[], bool] | None = None,
) -> ExperimentAnalysis:
    """Load one export directory without using or time-matching an Excel file."""
    source_path = Path(export_path).expanduser().resolve()
    if not source_path.is_dir():
        raise ValueError(f"轨迹实验目录不存在：{source_path}")
    started_at = parse_trajectory_export_time(source_path)
    if started_at is None:
        csv_paths = list(source_path.glob("cage_*/data/trajectory.csv"))
        if not csv_paths:
            raise ValueError(f"实验目录中没有 trajectory.csv：{source_path}")
        started_at = datetime.fromtimestamp(
            min(csv_path.stat().st_mtime for csv_path in csv_paths)
        )
    return _load_trajectory_csv_export(
        source_path,
        source_path,
        started_at,
        interruption_requested=interruption_requested,
    )


def load_experiment_data(
    excel_path: str | Path,
    trajectory_root: str | Path | None = None,
    interruption_requested: Callable[[], bool] | None = None,
) -> ExperimentAnalysis:
    source_path = Path(excel_path).expanduser().resolve()
    workbook_error: Exception | None = None
    try:
        workbook_analysis = load_experiment_workbook(
            source_path,
            interruption_requested=interruption_requested,
        )
        if any(channel.valid_rows for channel in workbook_analysis.channels.values()):
            return workbook_analysis
        workbook_error = ValueError("Excel轨迹工作表没有有效坐标")
    except InterruptedError:
        raise
    except Exception as error:
        workbook_error = error

    if interruption_requested is not None and interruption_requested():
        raise InterruptedError("实验数据读取已取消")
    if trajectory_root is None:
        raise workbook_error or ValueError("Excel中没有可分析的轨迹数据")

    started_at = parse_experiment_time(source_path)
    if started_at is None:
        started_at = datetime.fromtimestamp(source_path.stat().st_mtime)
    export_path = find_matching_trajectory_export(trajectory_root, started_at)
    if export_path is None:
        detail = str(workbook_error) if workbook_error is not None else "Excel中没有有效轨迹"
        raise ValueError(f"{detail}；未找到时间匹配的轨迹目录")
    return _load_trajectory_csv_export(
        source_path,
        export_path,
        started_at,
        interruption_requested=interruption_requested,
    )


def aggregate_distance(channel: ChannelTrajectory, bin_seconds: float) -> tuple[np.ndarray, np.ndarray]:
    if channel.total_rows == 0 or bin_seconds <= 0:
        return np.asarray([], dtype=float), np.asarray([], dtype=float)
    bin_indexes = np.floor(channel.elapsed_seconds / bin_seconds).astype(int)
    distances = np.bincount(
        bin_indexes,
        weights=channel.distance_steps_mm,
        minlength=int(bin_indexes.max()) + 1,
    )
    return np.arange(distances.size, dtype=float) * bin_seconds, distances


def trajectory_plot_arrays(channel: ChannelTrajectory) -> tuple[np.ndarray, np.ndarray]:
    if channel.total_rows == 0:
        return np.asarray([], dtype=float), np.asarray([], dtype=float)
    x_values = channel.x_mm.copy()
    y_values = channel.y_mm.copy()
    x_values[~channel.valid] = np.nan
    y_values[~channel.valid] = np.nan
    if channel.total_rows > 1:
        gaps = np.diff(channel.times) > channel.gap_limit_seconds
        break_indexes = np.flatnonzero(gaps) + 1
        x_values[break_indexes] = np.nan
        y_values[break_indexes] = np.nan
    return x_values, y_values


def sleep_state_matrix(
    analysis: ExperimentAnalysis,
    movement_threshold_mm: float,
    segment_seconds: float = 10.0,
    required_segments: int = 4,
    short_missing_seconds: float = 5.0,
) -> tuple[np.ndarray, np.ndarray]:
    duration = max(analysis.duration_seconds, segment_seconds)
    segment_count = max(1, int(math.ceil(duration / segment_seconds)))
    states = np.full((8, segment_count), np.nan, dtype=float)

    for channel_number in range(1, 9):
        channel = analysis.channels[channel_number]
        if channel.total_rows == 0:
            continue
        still_streak = 0
        for segment_index in range(segment_count):
            segment_start = segment_index * segment_seconds
            segment_end = segment_start + segment_seconds
            in_segment = (
                (channel.elapsed_seconds >= segment_start)
                & (channel.elapsed_seconds < segment_end)
            )
            observed_indexes = np.flatnonzero(in_segment)
            if observed_indexes.size == 0:
                still_streak = 0
                states[channel_number - 1, segment_index] = 0.0
                continue

            valid_indexes = observed_indexes[channel.valid[observed_indexes]]
            if valid_indexes.size == 0:
                is_still = True
            else:
                x_values = channel.x_mm[valid_indexes]
                y_values = channel.y_mm[valid_indexes]
                center_x = float(np.median(x_values))
                center_y = float(np.median(y_values))
                radius = np.hypot(x_values - center_x, y_values - center_y)
                inside_radius = bool(np.max(radius, initial=0.0) <= movement_threshold_mm)

                valid_elapsed = channel.elapsed_seconds[valid_indexes]
                boundary_times = np.concatenate(
                    ([segment_start], valid_elapsed, [segment_end])
                )
                longest_missing = float(np.max(np.diff(boundary_times), initial=0.0))
                is_still = inside_radius and longest_missing <= short_missing_seconds

            if is_still:
                still_streak += 1
            else:
                still_streak = 0
            states[channel_number - 1, segment_index] = (
                1.0 if still_streak >= required_segments else 0.0
            )

    time_minutes = np.arange(segment_count + 1, dtype=float) * segment_seconds / 60.0
    return states, time_minutes
